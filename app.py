from flask import Flask, render_template, request, send_file, after_this_request
import pandas as pd
import matplotlib.pyplot as plt
import os
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from io import BytesIO


app = Flask(__name__)

# Directorio local para almacenar archivos temporalmente
UPLOAD_FOLDER = 'uploads/'
PROCESSED_FOLDER = 'processed/'

# Crear carpetas si no existen
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

# Ruta principal
@app.route('/')
def index():
    return render_template('index.html')

# Ruta para subir y limpiar archivos
@app.route('/clean', methods=['POST'])
def upload_file():
    file = request.files['file']
    if file:
        # Guardar el archivo subido en el directorio local
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)
        print(f"Archivo {file.filename} guardado localmente en {file_path}.")

        # Procesar el archivo cargado
        processed_file_path = process_file(file_path)

        # Enviar el archivo procesado para su descarga
        return send_file(processed_file_path, as_attachment=True)
    return "No se subió ningún archivo"
def process_file(file_path):
    # Cargar el archivo Excel
    wb = load_workbook(file_path, data_only=True)
    cleaned_df = clean_and_prepare_excel(wb)

    # Guardar el DataFrame limpio en el directorio local
    processed_file_path = os.path.join(PROCESSED_FOLDER, 'processed_' + os.path.basename(file_path))
    cleaned_df.to_excel(processed_file_path, index=False, engine='openpyxl')
    print(f"Archivo procesado guardado como {processed_file_path}.")

    # Eliminar el archivo original después del procesamiento
    os.remove(file_path)
    print(f"Archivo original {file_path} eliminado.")
    download_file(processed_file_path)
    return processed_file_path

# Función para limpiar y preparar el df
def clean_and_prepare_excel(workbook):
    sheet = workbook.active

    # Descombinar celdas y copiar el valor solo a la primera celda
    merged_cells_copy = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_cells_copy:
        merged_cell_value = sheet[merged_cell_range.coord.split(':')[0]].value
        sheet.unmerge_cells(str(merged_cell_range))
        first_cell = True
        for row in sheet[merged_cell_range.coord]:
            for cell in row:
                if first_cell:
                    cell.value = merged_cell_value  # Mantener el valor solo en la primera celda
                    first_cell = False
                else:
                    cell.value = None  # Dejar las demás celdas vacías

    # Guardar el archivo descombinado en memoria
    temp_file = BytesIO()
    workbook.save(temp_file)
    temp_file.seek(0)

    # Leer el archivo temporal en un DataFrame de pandas para más limpieza
    df = pd.read_excel(temp_file)

    # Limpieza básica de datos
    df.drop_duplicates(inplace=True)  # Eliminar duplicados
    df.dropna(how='all', inplace=True)  # Eliminar filas completamente vacías
    df.fillna('', inplace=True)  # Reemplazar NaN con una cadena vacía

    # Normalizar formatos de números y fechas
    def is_date_string(s):
        date_patterns = [
            r"^\d{4}-\d{2}-\d{2}$",  # Formato yyyy-mm-dd
            r"^\d{2}/\d{2}/\d{4}$",  # Formato dd/mm/yyyy
            r"^\d{2}-\d{2}-\d{4}$",  # Formato dd-mm-yyyy
            r"^\d{2} \w{3,9} \d{4}$",  # Formato dd Month yyyy
        ]
        if isinstance(s, str):
            return any(re.match(pattern, s.strip()) for pattern in date_patterns)
        return False

    # Detectar y convertir columnas de fecha
    for col in df.columns:
        if df[col].dtype == 'object':  # Solo procesa columnas de texto
            date_matches = df[col].apply(is_date_string)
            if date_matches.sum() / len(df) > 0.7:  # Ajustar a más del 70% parecen fechas
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.date  # Mantener solo la fecha
                except Exception as e:
                    print(f"No se pudo convertir la columna {col} a fecha: {e}")

        # Convertir columnas numéricas a float
        if df[col].dtype == 'O':
            try:
                df[col] = pd.to_numeric(df[col], errors='ignore')
            except:
                pass
    return df

# Combinación de archivos
@app.route('/merge', methods=['POST'])
def merge_files():
    files = request.files.getlist('files')  # Permitir múltiples archivos
    if files:
        # Leer los archivos en memoria como DataFrames
        dataframes = [pd.read_excel(file) for file in files]

        # Combinar todos los DataFrames en uno solo
        merged_df = pd.concat(dataframes, ignore_index=True)

        # Nombre del archivo combinado
        merged_file_name = "merged_file.xlsx"
        merged_file_path = os.path.join(PROCESSED_FOLDER, merged_file_name)

        # Guardar el DataFrame combinado en un archivo Excel en el almacenamiento local
        merged_df.to_excel(merged_file_path, index=False, engine='openpyxl')
        print(f"Archivo combinado guardado como {merged_file_path}.")
        
        # Usamos after_this_request para ejecutar una acción después de que se complete la solicitud
        @after_this_request
        def remove_file(response):
            try:
                os.remove(merged_file_path)
                print(f"Archivo convertido {merged_file_path} eliminado después de la descarga.")
            except Exception as e:
                print(f"Error al intentar eliminar el archivo: {e}")
            return response

        # Proporcionar un enlace de descarga para el archivo combinado
        return send_file(merged_file_path, as_attachment=True, download_name=merged_file_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return "No se subieron archivos"

@app.route('/convert_format', methods=['POST'])
def convert_format():
    file = request.files['file']
    conversion_type = request.form['conversion_type']
    if file:
        # Inicializar variables para guardar el archivo convertido
        converted_file_path = None
        converted_filename = None

        # Leer el archivo en un DataFrame en memoria y convertir según el tipo
        if conversion_type == 'excel_to_csv':
            df = pd.read_excel(file)
            # Nombre del archivo convertido
            converted_filename = os.path.splitext(file.filename)[0] + ".csv"
            converted_file_path = os.path.join(PROCESSED_FOLDER, converted_filename)
            # Guardar como CSV localmente
            df.to_csv(converted_file_path, index=False)
            print(f"Archivo convertido guardado como {converted_file_path}.")

        elif conversion_type == 'csv_to_excel':
            df = pd.read_csv(file)
            # Nombre del archivo convertido
            converted_filename = os.path.splitext(file.filename)[0] + ".xlsx"
            converted_file_path = os.path.join(PROCESSED_FOLDER, converted_filename)
            # Guardar como Excel localmente
            df.to_excel(converted_file_path, index=False, engine='openpyxl')
            print(f"Archivo convertido guardado como {converted_file_path}.")

        else:
            return "Tipo de conversión no soportado."

        # Usamos after_this_request para ejecutar una acción después de que se complete la solicitud
        @after_this_request
        def remove_file(response):
            try:
                os.remove(converted_file_path)
                print(f"Archivo convertido {converted_file_path} eliminado después de la descarga.")
            except Exception as e:
                print(f"Error al intentar eliminar el archivo: {e}")
            return response

        # Proporcionar el archivo convertido para su descarga
        return send_file(converted_file_path, as_attachment=True, download_name=converted_filename, mimetype='text/csv' if conversion_type == 'excel_to_csv' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return "No se subió ningún archivo"

def download_file(filename):
    processed_file_path = os.path.join(PROCESSED_FOLDER, filename)

    if os.path.exists(processed_file_path):
        # Usamos after_this_request para ejecutar una acción después de que se complete la solicitud
        @after_this_request
        def remove_file(response):
            try:
                os.remove(processed_file_path)
                print(f"Archivo procesado {processed_file_path} eliminado después de la descarga.")
            except Exception as e:
                print(f"Error al intentar eliminar el archivo: {e}")
            return response

        # Enviamos el archivo para su descarga
        return send_file(processed_file_path, as_attachment=True)
    else:
        return "El archivo no existe", 404
        
if __name__ == "__main__":
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))